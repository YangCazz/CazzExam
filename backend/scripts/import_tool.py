import sys, os, json
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))


def make_template(path: str):
    """生成 Excel 导入模板（含表头与示例行）"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "题目"
    ws.append(["题型(choice/case/essay)", "科目(1综合/2案例/3论文)", "题干",
               "选项A", "选项B", "选项C", "选项D", "答案", "解析",
               "难度1-5", "年份", "来源(real/self/ai)", "知识点ID(逗号分隔)"])
    ws.append(["choice", 1, "以下哪种架构风格以数据为中心？",
               "调用/返回风格", "数据流风格", "仓库风格", "解释器风格", "C",
               "仓库风格以数据为中心。", 3, 2023, "real", "3,7"])
    ws.append(["case", 2, "某系统要求高并发低延迟，请识别其质量属性并给出架构方案。",
               "", "", "", "", "要点：性能/可用性；方案：微服务+缓存+限流……", "", 4, 2024, "real", "8,10"])
    wb.save(path)
    print("template saved:", path)


def xlsx2json(src: str, dst: str):
    from openpyxl import load_workbook
    wb = load_workbook(src)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    def _v(col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in rows[1:]:
        stem = _v("题干")
        if not stem:
            continue
        opts = [_v("选项A"), _v("选项B"), _v("选项C"), _v("选项D")]
        opts = [o for o in opts if o]
        kp_raw = _v("知识点ID(逗号分隔)")
        kp = [int(x.strip()) for x in str(kp_raw or "").split(",") if str(x).strip().isdigit()]
        year = _v("年份")
        qtype = _v("题型(choice/case/essay)")
        source_type = _v("来源(real/self/ai)")
        difficulty = _v("难度1-5")
        subject = _v("科目(1综合/2案例/3论文)")
        q = {
            "qtype": qtype.strip() if qtype else "choice",
            "subject": int(subject) if subject else 1,
            "stem": stem,
            "options": opts,
            "answer": _v("答案") or "",
            "analysis": _v("解析") or "",
            "difficulty": int(difficulty) if difficulty else 3,
            "source_type": source_type.strip() if source_type else "real",
            "source_year": int(year) if year else None,
            "knowledge_ids": kp,
        }
        out.append(q)
    with open(dst, "w", encoding="utf-8") as f:
        json.dump({"questions": out}, f, ensure_ascii=False, indent=1)
    print("converted", len(out), "questions ->", dst)


if __name__ == "__main__":
    cmd = sys.argv[1]
    if cmd == "template":
        make_template(sys.argv[2] if len(sys.argv) > 2 else "import_template.xlsx")
    elif cmd == "xlsx2json":
        xlsx2json(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "questions.json")
    else:
        print("usage: python import_tool.py template [out.xlsx] | xlsx2json <in.xlsx> [out.json]")
